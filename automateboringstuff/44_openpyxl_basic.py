import openpyxl
import os

os.chdir('/users/markli/Downloads/')

workbook = openpyxl.load_workbook('Book1.xlsx')

# print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')

# print(workbook.get_sheet_names())

cell = sheet['A1']
print(cell.value)

print(str(sheet['A1'].value))


#Another way to get the cell
cell = sheet.cell(row=1, column=2)  # It is the same with sheet['B1']


for i in range(1, 8):
    print(i, sheet.cell(row = i, column=2).value)

