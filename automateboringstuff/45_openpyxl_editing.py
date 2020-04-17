import openpyxl, os

wb = openpyxl.Workbook()  # create a blank workbook
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet')

print(sheet['A1'] == None)

sheet['A1'] = 200

sheet['A2'] = 'Hello'  # Now all the information is in our RAM

sheet2 = wb.create_sheet(title='sheeting')

sheet2['B2'] = 'testing'

wb.create_sheet(index=0, title='My first index sheet')  # This new sheet will be listed at the first place of the sheets positons

os.chdir('/users/markli/Downloads/')
wb.save('test.xlsx')



